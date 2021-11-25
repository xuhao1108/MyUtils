#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2021/7/22 16:43
# @Author : 闫旭浩
# @Email : 874591940@qq.com
# @desc : Excel的读写
import os
import xlrd
import openpyxl
import requests
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO

__all__ = ['YxhReadExcel', 'YxhWriteExcel']


class Base(object):
    @staticmethod
    def get_row_position(row):
        """
        获取行的表示方法
        :param row: 行下标
        :return:
        """
        return row + 1

    @staticmethod
    def get_col_position(col):
        """
        获取列的表示方法
        :param col: 列下标
        :return:
        """
        return xlrd.colname(col)

    @staticmethod
    def get_col_number(col):
        """
        获取列的表示方法
        :param col: 列英文字母
        :return:
        """
        number = 0
        for index, c in enumerate(col.lower()):
            number += (ord(c) - 96) * (26 ** (len(col) - 1 - index))
        return number - 1

    @staticmethod
    def get_cell_position(row, col):
        """
        获取单元格的表示方法
        :param row: 行下标
        :param col: 列下标
        :return:
        """
        return xlrd.cellname(row, col)

    @staticmethod
    def get_image_obj(image_path=None, image_url=None):
        """
        获取写入到excel的图片对象
        :param image_path: 图片路径
        :param image_url: 图片URL路径
        :return:
        """
        try:
            if image_url:
                response = requests.get(image_url)
                image_path = PILImage.open(BytesIO(response.content))
            return Image(image_path)
        except:
            return image_path


class YxhReadExcel(Base):
    def __init__(self, excel_path, sheet_index=None, sheet_name=None):
        self.excel = xlrd.open_workbook(excel_path)
        if sheet_name:
            self.sheet = self.excel.sheet_by_name(sheet_name)
        else:
            sheet_index = sheet_index if sheet_index else 0
            self.sheet = self.excel.sheet_by_index(sheet_index)

    def get_sheet_names(self):
        """
        获取所有sheet的名称
        :return:
        """
        return self.excel.sheet_names()

    def get_current_sheet_name(self):
        """
        获取当前sheet的名称
        :return:
        """
        return self.sheet.name

    def set_sheet_by_name(self, name):
        """
        切换到指定工作簿
        :param name: sheet名称
        :return:
        """
        self.sheet = self.excel.sheet_by_name(name)

    def get_row(self):
        """
        获取有效行数
        :return:
        """
        return self.sheet.nrows

    def get_col(self):
        """
        获取有效列数
        :return:
        """
        return self.sheet.ncols

    def read_cell(self, row, col):
        """
        读取单元格
        :param row: int 从0开始，第一行即为0...
        :param col: int 从0开始，第一列即为0...
        :return: object | None 单元格数据，None表示没有数据
        """
        try:
            return self.sheet.cell_value(row, col)
        except:
            return None

    def read_row(self, row=0):
        """
        读取行
        :param row: int 从0开始，第一行即为0...
        :return: list 行数据
        """
        return self.sheet.row_values(row)

    def read_col(self, col=0):
        """
        读取列
        :param col: int 从0开始，第一列即为0...
        :return:
        """
        return self.sheet.col_values(col)

    def read_range(self, start_row, start_col, stop_row, stop_col):
        """
        读取指定区域
        :param start_row: 起始行
        :param start_col: 起始列
        :param stop_row: 结束行
        :param stop_col: 结束列
        :return:
        """
        data = []
        # 依次读取行
        for row in range(start_row, stop_row):
            row_data = []
            # 依次读取行上的各单元格数据
            for col in range(start_col, stop_col):
                # 读取单元格
                row_data.append(self.read_cell(row, col))
            # 添加到行数据中
            data.append(row_data)
        return data

    def read_all_data(self, row_header=True, col_header=False):
        """
        获取excel所有数据，并按照
        :param row_header: 行首为标题
        :param col_header: 列首尾标题
        :return:
        """
        data = []
        rows = self.get_row()
        cols = self.get_col()
        if row_header:
            headers = self.read_row()
            for row in range(1, rows):
                row_data = self.read_row(row)
                data.append(dict(map(lambda x, y: [x, y], headers, row_data)))
        elif col_header:
            headers = self.read_col()
            for col in range(1, cols):
                col_data = self.read_col(col)
                data.append(dict(map(lambda x, y: [x, y], headers, col_data)))
        else:
            data = self.read_range(0, 0, rows, cols)
        return data


class YxhWriteExcel(Base):
    def __init__(self, excel_path, save_path=None, sheet_name=None, sheet_index=None, always_save=True):
        """
        初始化
        :param excel_path: excel_path路径
        :param save_path: excel_path保存路径
        :param sheet_name: 工作簿名称
        :param sheet_index: 工作簿下标
        :param always_save: True | False 实时保存
        """
        if os.path.exists(excel_path):
            self.excel = openpyxl.load_workbook(excel_path)
        else:
            self.excel = openpyxl.Workbook()
        if sheet_name:
            self.sheet = self.excel[sheet_name]
        else:
            self.sheet = self.excel[self.excel.sheetnames[sheet_index or 0]]
        self.save_path = save_path if save_path else excel_path
        self.always_save = always_save

    def get_sheet_names(self):
        """
        获取所有sheet名称
        :return:
        """
        return self.excel.sheetnames

    def get_current_sheet_name(self):
        """
        获取当前sheet的名称
        :return:
        """
        return self.sheet.name

    def create_sheet(self, name, index=None):
        """
        创建sheet
        :param name: sheet名称
        :param index: sheet下标
        :return:
        """
        self.excel.create_sheet(name, index=index)

    def get_sheet(self, sheet_name=None, sheet_index=None):
        """
        获取指定sheet
        :param sheet_name: sheet名称
        :param sheet_index: sheet下标
        :return:
        """
        if sheet_name:
            return self.excel[sheet_name]
        if sheet_index:
            return self.excel[self.excel.sheetnames[sheet_index]]

    def set_sheet(self, sheet_name=None, sheet_index=None):
        """
        设置当前sheet
        :param sheet_name: sheet名称
        :param sheet_index: sheet下标
        :return:
        """
        obj = self.get_sheet(sheet_name, sheet_index)
        if obj:
            self.sheet = obj

    def append_data(self, data, save_flag=True):
        """
        在末尾追加一行数据
        :param data: 追加的数据
        :param save_flag: True | False 是否立刻保存
        :return:
        """
        self.sheet.append(data)
        if self.always_save or save_flag:
            self.save_to_excel()

    def write_cell(self, data, row=0, col=0, cell_size=None, image_size=None, save_flag=True):
        """
        写入单元格
        :param data: 写入数据
        :param row: 行，从0开始
        :param col: 列，从0开始
        :param cell_size: {'width':0, 'height':0} 单元格大小
        :param image_size: {'width':0, 'height':0} 图片大小
        :param save_flag: True | False 是否立刻保存
        :return:
        """
        # 设置单元格大小
        if image_size:
            try:
                self.set_col_width(float(cell_size['width']), col)
            except:
                pass
            try:
                self.set_row_height(float(cell_size['height']), row)
            except:
                pass
        position = self.get_cell_position(row, col)
        if isinstance(data, Image):
            # 设置图片大小
            if image_size:
                try:
                    data.width = float(image_size['width'])
                except:
                    pass
                try:
                    data.height = float(image_size['height'])
                except:
                    pass
            # 插入图片
            self.sheet.add_image(data, position)
        else:
            # 插入数据
            self.sheet[position] = data
        if self.always_save or save_flag:
            self.save_to_excel()

    def write_row(self, data, row=0, save_flag=True):
        """
        写入行
        :param data: list 写入数据
        :param row: 行，从0开始
        :param save_flag: True | False 是否立刻保存
        :return:
        """
        # 依次写入此行每个单元格
        for col, value in enumerate(data):
            self.write_cell(value, row, col)
        if self.always_save or save_flag:
            self.save_to_excel()

    def write_col(self, data, col=0, save_flag=True):
        """
        写入列
        :param data: list 写入数据
        :param col: 列，从0开始
        :param save_flag: True | False 是否立刻保存
        :return:
        """
        # 依次写入此列每个单元格
        for row, value in enumerate(data):
            self.write_cell(value, row, col)
        if self.always_save or save_flag:
            self.save_to_excel()

    def write_range(self, data, start_row=0, start_col=0, save_flag=True):
        """
        写入指定区域
        :param data: 写入数据
        :param start_row: 写入的起始行
        :param start_col: 写入的起始列
        :param save_flag: True | False 是否立刻保存
        :return:
        """
        for row, row_data in enumerate(data):
            for col, value in enumerate(row_data):
                self.write_cell(value, start_row + row, start_col + col)
        if self.always_save or save_flag:
            self.save_to_excel()

    def delete_cell(self, row=0, col=0, save_flag=True):
        pass

    def delete_row(self, row=0, save_flag=True):
        """
        删除行
        :param row: 行号
        :param save_flag: True | False 是否立刻保存
        :return:
        """
        self.sheet.delete_rows(self.get_row_position(row))
        if self.always_save or save_flag:
            self.save_to_excel()

    def delete_col(self, col=0, save_flag=True):
        """
        删除列
        :param col: 列号
        :param save_flag: True | False 是否立刻保存
        :return:
        """
        self.sheet.delete_cols(self.get_col_position(col))
        if self.always_save or save_flag:
            self.save_to_excel()

    def delete_range(self, start_row=0, start_col=0, save_flag=True):
        pass

    def set_row_height(self, height, row=0, save_flag=True):
        """
        设置行高
        :param height: 行高
        :param row: 行号
        :param save_flag: True | False 是否立刻保存
        :return:
        """
        self.sheet.row_dimensions[self.get_row_position(row)].height = height
        if self.always_save or save_flag:
            self.save_to_excel()

    def set_col_width(self, width, col=0, save_flag=True):
        """
        设置列宽
        :param width: 列宽
        :param col: 列号
        :param save_flag: True | False 是否立刻保存
        :return:
        """
        self.sheet.column_dimensions[self.get_col_position(col)].width = width
        if self.always_save or save_flag:
            self.save_to_excel()

    def merge_cell(self, start_row, start_col, stop_row, stop_col, save_flag=True):
        """
        合并单元格
        :param start_row: 开始行
        :param start_col: 开始列
        :param stop_row: 结束行
        :param stop_col: 结束列
        :param save_flag: True | False 是否立刻保存
        :return:
        """
        self.sheet.merge_cells('{}:{}'.format(self.get_cell_position(start_row, start_col), self.get_cell_position(stop_row, stop_col)))
        if self.always_save or save_flag:
            self.save_to_excel()

    def save_to_excel(self):
        """
        保存excel
        :return:
        """
        self.excel.save(self.save_path)


if __name__ == '__main__':
    pass
